"""
attendance.py – Excel-based attendance record management.

Handles:
- Creating / opening Attendance_Sheet_YYYY.xlsx files.
- Adding start / end times to the correct year-file and month-sheet.
- Calculating work time (with 1-hour break deduction when > 6 h).
- Sorting rows by date ascending.
- Writing a monthly total row when all days in the month are present.
"""

import calendar
import glob
import os
from datetime import date, datetime
from typing import Optional, Tuple

from openpyxl import Workbook, load_workbook

HEADERS = ["日付", "始業時間", "終業時間", "労働時間"]
DATE_FMT = "%Y/%m/%d"
TIME_FMT = "%H:%M"


class AttendanceManager:
    """Reads and writes attendance data to Excel files."""

    def __init__(self, config) -> None:
        self.config = config

    # ------------------------------------------------------------------
    # Path helpers
    # ------------------------------------------------------------------

    def get_file_path(self, year: Optional[int] = None) -> Optional[str]:
        """Return the full path of the Excel file for *year* (default: current year)."""
        if not self.config.folder_path:
            return None
        if year is None:
            year = datetime.now().year
        return os.path.join(
            self.config.folder_path, f"Attendance_Sheet_{year}.xlsx"
        )

    @staticmethod
    def get_sheet_name(month: Optional[int] = None) -> str:
        """Return the abbreviated month name used as the sheet name (e.g. 'Jan')."""
        if month is None:
            month = datetime.now().month
        return datetime(2000, month, 1).strftime("%b")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def record_start(self, dt: datetime) -> Optional[str]:
        """Write start time for *dt* to the Excel file.  Returns 'HH:MM' or None."""
        return self._write_time(dt, column=1)

    def record_end(self, dt: datetime) -> Optional[str]:
        """Write end time for *dt* to the Excel file.  Returns 'HH:MM' or None."""
        return self._write_time(dt, column=2)

    def get_today_times(self) -> Tuple[Optional[str], Optional[str]]:
        """Return (start_time, end_time) strings for today, or (None, None)."""
        now = datetime.now()
        file_path = self.get_file_path(now.year)
        if not file_path or not os.path.exists(file_path):
            return None, None

        sheet_name = self.get_sheet_name(now.month)
        try:
            wb = load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                return None, None
            ws = wb[sheet_name]
            date_str = now.strftime(DATE_FMT)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == date_str:
                    return row[1], row[2]
        except Exception as exc:
            print(f"[AttendanceManager] get_today_times error: {exc}")
        return None, None

    def fill_missing_end_time(self, target_date: date, end_dt: datetime) -> bool:
        """Fill a missing end time for *target_date* (date object) with *end_dt*.

        Returns True if the record was updated, False otherwise.
        """
        file_path = self.get_file_path(target_date.year)
        if not file_path or not os.path.exists(file_path):
            return False

        sheet_name = self.get_sheet_name(target_date.month)
        try:
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            data = self._read_data(ws)
            date_str = target_date.strftime(DATE_FMT)
            for row in data:
                if row[0] == date_str and row[1] and not row[2]:
                    end_str = end_dt.strftime(TIME_FMT)
                    row[2] = end_str
                    row[3] = self._calc_work_time(row[1], end_str)
                    self._flush(ws, data, target_date.year, target_date.month)
                    wb.save(file_path)
                    return True
        except Exception as exc:
            print(f"[AttendanceManager] fill_missing_end_time error: {exc}")
        return False

    def find_latest_missing_end_date(self) -> Optional[date]:
        """Scan all Attendance_Sheet_*.xlsx files and return the latest date that has
        a start time but no end time.  Returns None if no such date exists or the
        folder is not configured.
        """
        if not self.config.folder_path:
            return None

        pattern = os.path.join(self.config.folder_path, "Attendance_Sheet_*.xlsx")
        today = datetime.now().date()
        latest: Optional[date] = None

        for file_path in glob.glob(pattern):
            try:
                wb = load_workbook(file_path, data_only=True)
                try:
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and "合計" in str(row[0]):
                                continue
                            date_val = row[0]
                            start_val = row[1]
                            end_val = row[2]
                            if date_val and start_val and not end_val:
                                try:
                                    d = datetime.strptime(str(date_val).strip(), DATE_FMT).date()
                                except ValueError:
                                    continue
                                if d < today and (latest is None or d > latest):
                                    latest = d
                finally:
                    wb.close()
            except Exception as exc:
                print(f"[AttendanceManager] find_latest_missing_end_date error reading {file_path}: {exc}")

        return latest

    def check_previous_day(self) -> None:
        """Find the most recent workday with a missing end time and auto-fill it
        from Windows Event Log.

        Search covers all Attendance_Sheet_*.xlsx files so year/month boundaries
        are handled correctly.  The inferred end time is only accepted if it falls
        within 12:00–23:59 on the target date.
        """
        from windows_events import get_last_work_end_time

        target_date = self.find_latest_missing_end_date()
        if target_date is None:
            return

        end_time = get_last_work_end_time(target_date)
        if end_time is None:
            return

        # Only accept end times in the 12:00–23:59:59 window on the target date
        window_start = datetime(target_date.year, target_date.month, target_date.day, 12, 0)
        window_end = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)
        if not (window_start <= end_time <= window_end):
            return

        try:
            self.fill_missing_end_time(target_date, end_time)
        except Exception as exc:
            print(f"[AttendanceManager] check_previous_day error: {exc}")

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _write_time(self, dt: datetime, column: int) -> Optional[str]:
        """Internal: write start (column=1) or end (column=2) time."""
        if not self.config.folder_path:
            return None

        file_path = self.get_file_path(dt.year)
        sheet_name = self.get_sheet_name(dt.month)

        wb = self._open_or_create_workbook(file_path)
        ws = self._open_or_create_sheet(wb, sheet_name)

        date_str = dt.strftime(DATE_FMT)
        time_str = dt.strftime(TIME_FMT)

        data = self._read_data(ws)

        # Find existing row for today or create a new one
        target_row = None
        for row in data:
            if row[0] == date_str:
                target_row = row
                break
        if target_row is None:
            target_row = [date_str, None, None, None]
            data.append(target_row)

        target_row[column] = time_str

        # Recalculate work time whenever start or end changes
        target_row[3] = self._calc_work_time(target_row[1], target_row[2])

        self._flush(ws, data, dt.year, dt.month)
        wb.save(file_path)
        return time_str

    @staticmethod
    def _open_or_create_workbook(file_path: str) -> Workbook:
        if os.path.exists(file_path):
            return load_workbook(file_path)
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default empty sheet
        return wb

    @staticmethod
    def _open_or_create_sheet(wb: Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        # Basic column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        return ws

    @staticmethod
    def _read_data(ws) -> list:
        """Read data rows (skip header and total rows) as a list of lists."""
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and "合計" not in str(row[0]):
                rows.append(list(row))
        return rows

    @staticmethod
    def _calc_work_time(start_str: Optional[str], end_str: Optional[str]) -> Optional[str]:
        """Return work time as 'HH:MM', applying 1-hour break when > 6 hours."""
        if not start_str or not end_str:
            return None
        try:
            fmt = "%H:%M"
            start = datetime.strptime(str(start_str).strip(), fmt)
            end = datetime.strptime(str(end_str).strip(), fmt)
            minutes = (end - start).total_seconds() / 60
            if minutes <= 0:
                minutes += 24 * 60  # midnight crossing
            if minutes > 360:  # > 6 hours → deduct 1-hour break
                minutes -= 60
            h, m = divmod(int(minutes), 60)
            return f"{h:02d}:{m:02d}"
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _calc_total_time(data: list) -> str:
        """Sum all work times and return 'HH:MM'."""
        total_minutes = 0
        for row in data:
            if row[3]:
                parts = str(row[3]).split(":")
                try:
                    total_minutes += int(parts[0]) * 60 + int(parts[1])
                except (ValueError, IndexError):
                    pass
        h, m = divmod(total_minutes, 60)
        return f"{h:02d}:{m:02d}"

    def _flush(self, ws, data: list, year: int, month: int) -> None:
        """Write sorted data back to *ws*, adding a total row if month is full."""
        # Sort by date string (YYYY/MM/DD sorts lexicographically)
        data.sort(key=lambda r: str(r[0]) if r[0] else "")

        # Clear all existing data rows (keep header in row 1)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write sorted data
        for row in data:
            ws.append(row)

        # Add total row if every calendar day in the month has a record
        last_day = calendar.monthrange(year, month)[1]
        if len(data) == last_day:
            total = self._calc_total_time(data)
            ws.append(["合計", "", "", total])
